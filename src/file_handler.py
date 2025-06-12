import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# ? import os

def load_file(path):
    if path.endswith(".csv"):
        df = pd.read_csv(path, dtype=str)
    elif path.endswith((".xls", ".xlsx")):
        df = pd.read_excel(path, dtype=str)
    else:
        raise ValueError("Unsupported file format")

    # Strip whitespace from column headers
    df.columns = df.columns.str.strip()

    # Replace NaN with blank strings
    df = df.fillna("")

    # Strip whitespace from string cells
    df = df.astype(str).apply(lambda col: col.str.strip())

    return df

def save_file(df, output_path):
    # Save DataFrame to Excel first (unstyled)
    df.to_excel(output_path, index=False)

    # Load Excel file for styling
    wb = load_workbook(output_path)
    ws = wb.active

    yellow_fill = PatternFill(start_color="F8D12A", end_color="F8D12A", fill_type="solid")

    # Get column headers and map to Excel column index (1-based)
    headers = [cell.value for cell in ws[1]]
    col_name_to_index = {name: i + 1 for i, name in enumerate(headers)}

    # Identify the Validation_Errors column index
    error_col_index = col_name_to_index.get("Validation_Errors")

    for row_idx in range(2, ws.max_row + 1):  # skip header row
        error_cell = ws.cell(row=row_idx, column=error_col_index)
        if not error_cell.value:
            continue

        error_text = str(error_cell.value)
        for col_name in headers:
            if col_name == "Validation_Errors":
                continue
            if col_name in error_text:
                col_idx = col_name_to_index[col_name]
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    wb.save(output_path)

    # Return heatmap grid for consolidated output
    if "Validation_Errors" in df.columns:
        return generate_error_grid(df, grid_rows=10, grid_cols=8)
    return None

def generate_error_grid(df, grid_rows=10, grid_cols=8):
    # Build a boolean mask of where errors occur by column
    error_mask = pd.DataFrame(False, index=df.index, columns=df.columns[:-1])

    for row_idx, error_str in df["Validation_Errors"].astype(str).items():
        for col in df.columns[:-1]:  # skip the "Validation_Errors" column itself
            if col in error_str:
                error_mask.at[row_idx, col] = True

    n_rows, n_cols = error_mask.shape
    row_chunk = math.ceil(n_rows / grid_rows)
    col_chunk = math.ceil(n_cols / grid_cols)

    grid = np.zeros((grid_rows, grid_cols))

    for i in range(grid_rows):
        for j in range(grid_cols):
            r_start = i * row_chunk
            r_end = min((i + 1) * row_chunk, n_rows)
            c_start = j * col_chunk
            c_end = min((j + 1) * col_chunk, n_cols)

            block = error_mask.iloc[r_start:r_end, c_start:c_end]
            total_cells = block.size
            error_cells = block.values.sum()
            grid[i, j] = error_cells / total_cells if total_cells > 0 else 0

    return grid

def save_all_heatmaps(heatmap_dict, output_path):
    """Save all 20x2 heatmaps into a single Excel file, one sheet per role."""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for role_name, grid in heatmap_dict.items():
            if grid is not None:
                df_grid = pd.DataFrame(grid)
                safe_sheet_name = role_name[:31] 
                df_grid.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=False)